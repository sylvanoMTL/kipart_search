"""Tests for the BOM export engine (Story 2.1)."""

from __future__ import annotations

import csv
import sys
from pathlib import Path

import pytest

from kipart_search.core.bom_export import (
    JLCPCB_TEMPLATE,
    NEWBURY_TEMPLATE,
    PCBWAY_TEMPLATE,
    PRESET_TEMPLATES,
    BOMTemplate,
    export_bom,
    _natural_sort_key,
)
from kipart_search.core.models import (
    BoardComponent,
    detect_mount_type,
    extract_package_from_footprint,
)


# ── Fixtures ──

def _make_comp(ref: str, value: str, footprint: str, mpn: str = "",
               manufacturer: str = "") -> BoardComponent:
    extra = {}
    if manufacturer:
        extra["manufacturer"] = manufacturer
    return BoardComponent(
        reference=ref, value=value, footprint=footprint,
        mpn=mpn, extra_fields=extra,
    )


@pytest.fixture
def three_components() -> list[BoardComponent]:
    """Two resistors with same MPN + one cap."""
    return [
        _make_comp("R1", "10k", "Resistor_SMD:R_0805_2012Metric",
                    mpn="RC0805FR-0710KL", manufacturer="Yageo"),
        _make_comp("R2", "10k", "Resistor_SMD:R_0805_2012Metric",
                    mpn="RC0805FR-0710KL", manufacturer="Yageo"),
        _make_comp("C1", "100nF", "Capacitor_SMD:C_0402_1005Metric",
                    mpn="GRM155R71C104KA88D", manufacturer="Murata"),
    ]


# ── Template Tests ──

class TestPCBWayTemplate:
    def test_has_9_columns(self):
        assert len(PCBWAY_TEMPLATE.columns) == 9

    def test_column_headers(self):
        headers = [c.header for c in PCBWAY_TEMPLATE.columns]
        assert headers == [
            "Item #", "Designator", "Qty", "Manufacturer", "Mfg Part #",
            "Description / Value", "Package/Footprint", "Type",
            "Your Instructions / Notes",
        ]

    def test_file_format_xlsx(self):
        assert PCBWAY_TEMPLATE.file_format == "xlsx"

    def test_group_by_mpn(self):
        assert PCBWAY_TEMPLATE.group_by == "mpn"


# ── Grouping Tests ──

class TestGrouping:
    def test_grouping_by_mpn(self, three_components, tmp_path):
        out = tmp_path / "bom.xlsx"
        export_bom(three_components, PCBWAY_TEMPLATE, out)

        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        assert len(rows) == 2

        # R1+R2 grouped
        r_row = next(r for r in rows if "R1" in str(r[1]))
        assert r_row[1] == "R1,R2"
        assert r_row[2] == 2
        assert r_row[4] == "RC0805FR-0710KL"

        # C1 alone
        c_row = next(r for r in rows if "C1" in str(r[1]))
        assert c_row[1] == "C1"
        assert c_row[2] == 1


# ── Natural Sort Tests ──

class TestNaturalSort:
    def test_designator_natural_sort(self):
        refs = ["R1", "R10", "R2"]
        sorted_refs = sorted(refs, key=_natural_sort_key)
        assert sorted_refs == ["R1", "R2", "R10"]

    def test_mixed_prefixes(self):
        refs = ["C1", "C10", "C2", "R1"]
        sorted_refs = sorted(refs, key=_natural_sort_key)
        assert sorted_refs == ["C1", "C2", "C10", "R1"]


# ── Package Extraction Tests ──

class TestPackageExtraction:
    def test_0805(self):
        assert extract_package_from_footprint("Capacitor_SMD:C_0805_2012Metric") == "0805"

    def test_0402(self):
        assert extract_package_from_footprint("Resistor_SMD:R_0402_1005Metric") == "0402"

    def test_soic(self):
        assert extract_package_from_footprint("Package_SO:SOIC-8_3.9x4.9mm_P1.27mm") == "SOIC-8"

    def test_sot23(self):
        assert extract_package_from_footprint("Package_TO_SOT_SMD:SOT-23") == "SOT-23"

    def test_unknown(self):
        assert extract_package_from_footprint("Custom:MyFootprint") == ""


# ── SMD/THT Detection Tests ──

class TestMountTypeDetection:
    def test_resistor_smd(self):
        assert detect_mount_type("Resistor_SMD:R_0805_2012Metric") == "SMD"

    def test_package_dip_tht(self):
        assert detect_mount_type("Package_DIP:DIP-8_W7.62mm") == "THT"

    def test_package_so_smd(self):
        assert detect_mount_type("Package_SO:SOIC-8") == "SMD"

    def test_package_to_sot_tht(self):
        assert detect_mount_type("Package_TO_SOT_THT:TO-220-3") == "THT"

    def test_capacitor_tht(self):
        assert detect_mount_type("Capacitor_THT:C_Disc_D3.0mm_W2.0mm_P2.50mm") == "THT"

    def test_connector_pin_header_tht(self):
        assert detect_mount_type("Connector_PinHeader:PinHeader_1x04") == "THT"

    def test_default_smd(self):
        assert detect_mount_type("SomeUnknownLib:Part") == "SMD"


# ── Excel Output Tests ──

class TestExcelOutput:
    def test_excel_output(self, three_components, tmp_path):
        out = tmp_path / "bom.xlsx"
        result = export_bom(three_components, PCBWAY_TEMPLATE, out)

        assert result == out
        assert out.exists()

        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb.active

        # Check header row
        headers = [cell.value for cell in ws[1]]
        assert headers == [c.header for c in PCBWAY_TEMPLATE.columns]

        # Check data rows exist
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 2

    def test_item_numbers_sequential(self, three_components, tmp_path):
        out = tmp_path / "bom.xlsx"
        export_bom(three_components, PCBWAY_TEMPLATE, out)

        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        item_numbers = [r[0] for r in rows]
        assert item_numbers == [1, 2]

    def test_package_column_populated(self, three_components, tmp_path):
        out = tmp_path / "bom.xlsx"
        export_bom(three_components, PCBWAY_TEMPLATE, out)

        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        # Package is column index 6 (0-based)
        packages = [r[6] for r in rows]
        assert "0805" in packages
        assert "0402" in packages

    def test_type_column_populated(self, three_components, tmp_path):
        out = tmp_path / "bom.xlsx"
        export_bom(three_components, PCBWAY_TEMPLATE, out)

        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        # Type is column index 7
        types = [r[7] for r in rows]
        assert all(t == "SMD" for t in types)


# ── CSV Output Tests ──

class TestCSVOutput:
    def test_csv_output(self, three_components, tmp_path):
        csv_template = BOMTemplate(
            name="PCBWay-CSV", columns=PCBWAY_TEMPLATE.columns,
            file_format="csv",
        )
        out = tmp_path / "bom.csv"
        result = export_bom(three_components, csv_template, out)

        assert result == out
        assert out.exists()

        with open(out, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

        # Header + 2 data rows
        assert len(rows) == 3
        assert rows[0] == [c.header for c in PCBWAY_TEMPLATE.columns]


# ── Edge Cases ──

class TestEdgeCases:
    def test_empty_component_list(self, tmp_path):
        out = tmp_path / "empty.xlsx"
        export_bom([], PCBWAY_TEMPLATE, out)

        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb.active

        # Header only, no data rows
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 0

        # Header exists
        headers = [cell.value for cell in ws[1]]
        assert len(headers) == 9

    def test_components_with_no_mpn(self, tmp_path):
        comps = [
            _make_comp("R1", "10k", "Resistor_SMD:R_0805_2012Metric", mpn=""),
            _make_comp("R2", "10k", "Resistor_SMD:R_0805_2012Metric", mpn=""),
        ]
        out = tmp_path / "no_mpn.xlsx"
        export_bom(comps, PCBWAY_TEMPLATE, out)

        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        # Both grouped into one row (same empty MPN)
        assert len(rows) == 1
        assert rows[0][2] == 2  # qty
        assert rows[0][4] in ("", None)  # empty MPN (openpyxl returns None for empty cells)


# ── No GUI Import Test ──

class TestNoGUIImports:
    def test_no_pyside6_in_modules(self):
        # Force fresh import check
        import importlib
        importlib.import_module("kipart_search.core.bom_export")
        pyside_modules = [k for k in sys.modules if k.startswith("PySide6")]
        # PySide6 may be loaded by other tests, so check bom_export source instead
        source_path = Path(__file__).parent.parent.parent / "src" / "kipart_search" / "core" / "bom_export.py"
        content = source_path.read_text()
        assert "PySide6" not in content
        assert "from kipart_search.gui" not in content


# ── JLCPCB Template Tests ──

class TestJLCPCBTemplate:
    def test_has_4_columns(self):
        assert len(JLCPCB_TEMPLATE.columns) == 4

    def test_column_headers(self):
        headers = [c.header for c in JLCPCB_TEMPLATE.columns]
        assert headers == ["Comment", "Designator", "Footprint", "JLCPCB Part #"]

    def test_file_format_csv(self):
        assert JLCPCB_TEMPLATE.file_format == "csv"

    def test_export_with_lcsc_part(self, tmp_path):
        comps = [
            BoardComponent(
                reference="C1", value="100nF",
                footprint="Capacitor_SMD:C_0402_1005Metric",
                mpn="GRM155R71C104KA88D",
                extra_fields={"manufacturer": "Murata", "LCSC Part": "C12345"},
            ),
        ]
        out = tmp_path / "jlcpcb.csv"
        export_bom(comps, JLCPCB_TEMPLATE, out)

        with open(out, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

        assert rows[0] == ["Comment", "Designator", "Footprint", "JLCPCB Part #"]
        assert rows[1][0] == "100nF"       # Comment = value
        assert rows[1][1] == "C1"          # Designator
        assert rows[1][2] == "0402"        # Footprint = extracted package
        assert rows[1][3] == "C12345"      # JLCPCB Part #

    def test_export_without_lcsc_part(self, tmp_path):
        comps = [
            BoardComponent(
                reference="R1", value="10k",
                footprint="Resistor_SMD:R_0805_2012Metric",
                mpn="RC0805FR-0710KL",
                extra_fields={"manufacturer": "Yageo"},
            ),
        ]
        out = tmp_path / "jlcpcb_no_lcsc.csv"
        export_bom(comps, JLCPCB_TEMPLATE, out)

        with open(out, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

        assert rows[1][3] == ""  # Empty JLCPCB Part #


# ── Newbury Template Tests ──

class TestNewburyTemplate:
    def test_has_9_columns(self):
        assert len(NEWBURY_TEMPLATE.columns) == 9

    def test_column_headers(self):
        headers = [c.header for c in NEWBURY_TEMPLATE.columns]
        assert headers == [
            "Item#", "Description", "Quantity", "Manufacturer Name",
            "Manufacturer Part Number", "Supplier Name", "Supplier Part Number",
            "Designator", "Notes",
        ]

    def test_file_format_xlsx(self):
        assert NEWBURY_TEMPLATE.file_format == "xlsx"

    def test_export_with_supplier_fields(self, tmp_path):
        comps = [
            BoardComponent(
                reference="U1", value="STM32F103",
                footprint="Package_QFP:LQFP-48_7x7mm_P0.5mm",
                mpn="STM32F103C8T6",
                extra_fields={
                    "manufacturer": "STMicroelectronics",
                    "supplier_name": "Farnell",
                    "supplier_pn": "123-456",
                    "description": "MCU ARM Cortex-M3 64KB Flash",
                },
            ),
        ]
        out = tmp_path / "newbury.xlsx"
        export_bom(comps, NEWBURY_TEMPLATE, out)

        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        assert len(rows) == 1
        row = rows[0]
        assert row[0] == 1                              # Item#
        assert row[1] == "MCU ARM Cortex-M3 64KB Flash"  # Description
        assert row[2] == 1                              # Quantity
        assert row[3] == "STMicroelectronics"            # Manufacturer Name
        assert row[4] == "STM32F103C8T6"                 # Manufacturer Part Number
        assert row[5] == "Farnell"                       # Supplier Name
        assert row[6] == "123-456"                       # Supplier Part Number
        assert row[7] == "U1"                            # Designator
        assert row[8] in ("", None)                      # Notes

    def test_description_field(self, tmp_path):
        # No description in extra_fields → falls back to value
        comps = [
            BoardComponent(
                reference="R1", value="10k",
                footprint="Resistor_SMD:R_0805_2012Metric",
                mpn="RC0805FR-0710KL",
                extra_fields={"manufacturer": "Yageo"},
            ),
        ]
        out = tmp_path / "newbury_desc.xlsx"
        export_bom(comps, NEWBURY_TEMPLATE, out)

        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        assert rows[0][1] == "10k"  # Description falls back to value


# ── Preset Templates Tests ──

class TestPresetTemplates:
    def test_preset_list_contains_all(self):
        assert len(PRESET_TEMPLATES) == 3
        assert PRESET_TEMPLATES[0] is PCBWAY_TEMPLATE
        assert PRESET_TEMPLATES[1] is JLCPCB_TEMPLATE
        assert PRESET_TEMPLATES[2] is NEWBURY_TEMPLATE

    def test_preset_names(self):
        names = [t.name for t in PRESET_TEMPLATES]
        assert names == ["PCBWay", "JLCPCB", "Newbury Electronics"]


# ── CSV Export All Templates Tests ──

class TestCSVExportAllTemplates:
    def test_jlcpcb_csv_output(self, tmp_path):
        comps = [
            BoardComponent(
                reference="C1", value="100nF",
                footprint="Capacitor_SMD:C_0402_1005Metric",
                mpn="GRM155R71C104KA88D",
                extra_fields={"manufacturer": "Murata", "LCSC Part": "C12345"},
            ),
            BoardComponent(
                reference="C2", value="100nF",
                footprint="Capacitor_SMD:C_0402_1005Metric",
                mpn="GRM155R71C104KA88D",
                extra_fields={"manufacturer": "Murata", "LCSC Part": "C12345"},
            ),
        ]
        out = tmp_path / "jlcpcb_full.csv"
        export_bom(comps, JLCPCB_TEMPLATE, out)

        with open(out, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

        assert rows[0] == ["Comment", "Designator", "Footprint", "JLCPCB Part #"]
        assert len(rows) == 2  # header + 1 grouped row
        assert rows[1][0] == "100nF"
        assert rows[1][1] == "C1,C2"
        assert rows[1][3] == "C12345"

    def test_newbury_csv_output(self, tmp_path):
        comps = [
            BoardComponent(
                reference="R1", value="10k",
                footprint="Resistor_SMD:R_0805_2012Metric",
                mpn="RC0805FR-0710KL",
                extra_fields={"manufacturer": "Yageo", "description": "RES 10K 1% 0805"},
            ),
        ]
        csv_newbury = BOMTemplate(
            name=NEWBURY_TEMPLATE.name,
            columns=NEWBURY_TEMPLATE.columns,
            file_format="csv",
        )
        out = tmp_path / "newbury.csv"
        export_bom(comps, csv_newbury, out)

        with open(out, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

        expected_headers = [c.header for c in NEWBURY_TEMPLATE.columns]
        assert rows[0] == expected_headers
        assert rows[1][1] == "RES 10K 1% 0805"  # Description from extra_fields
        assert rows[1][4] == "RC0805FR-0710KL"   # MPN
